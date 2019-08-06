from googletrans import Translator
from openpyxl import Workbook,load_workbook
import sys 
import os 

arg1 = sys.argv[1]
arg2 = sys.argv[2]
translator = Translator()
if os.path.exists(arg1):
	outerlist = []
	wk = load_workbook(arg1)
	ws = wk.active
	for row in ws.iter_rows():
		for cell in row:
			cell.value= translator.translate(str(cell.value)).text
			# print(cell.value)
		
	wk.save(str(arg2)+".xlsx")
